from dataclasses import dataclass
from openpyxl import load_workbook


@dataclass
class WarehouseOrder:
    order_number: str
    trailer: str
    location: str
    product: str
    pallets: int
    status: str = "Available"
    trip: str = ""


class Warehouse:

    def __init__(self, workbook_path):
        self.workbook_path = workbook_path
        self.orders: list[WarehouseOrder] = []
        self.doors = []
        self.trailers = []

    def load(self) -> bool:
        try:
            workbook = load_workbook(
                self.workbook_path,
                data_only=True
            )

            worksheet = workbook["WarehouseData"]

        except Exception as error:
            print(f"Warehouse load error: {error}")
            return False

        # Prevent duplicate records when reloading.
        self.orders.clear()

        for row in worksheet.iter_rows(
            min_row=2,
            values_only=True
        ):
            if not row or row[0] is None:
                continue

            try:
                pallets = int(row[4] or 0)
            except (TypeError, ValueError):
                pallets = 0

            order = WarehouseOrder(
                order_number=str(row[0] or "").strip(),
                trailer=str(row[1] or "").strip(),
                location=str(row[2] or "").strip(),
                product=str(row[3] or "").strip(),
                pallets=pallets,
                status=str(row[5] or "Available").strip(),
                trip=str(row[6] or "").strip()
                if len(row) > 6 else ""
            )

            self.orders.append(order)

        return True

    def get_orders(self) -> list[WarehouseOrder]:
        return self.orders.copy()

    def get_location(self, location: str) -> list[WarehouseOrder]:
        return [
            order
            for order in self.orders
            if order.location.lower() == location.lower()
        ]

    def get_locations(self) -> list[str]:
        return sorted({
            order.location
            for order in self.orders
            if order.location
        })

    def find_order(
        self,
        order_number: str
    ) -> WarehouseOrder | None:

        order_number = str(order_number).strip()

        for order in self.orders:
            if order.order_number == order_number:
                return order

        return None

    def reload(self) -> bool:
        return self.load()