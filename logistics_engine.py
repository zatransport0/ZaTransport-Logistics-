"""
ZaTransport Logistics Engine

Reads the Orders worksheet
Detects Zones
Loads every order into memory
"""

from openpyxl import load_workbook
from dataclasses import dataclass
from models import Route
from warehouse import Warehouse

@dataclass
class Order:

    zone: int

    shipping_company: str

    origin: str

    order_number: str

    product: str

    pallets: int

    customer: str

    destination: str

    warehouse_location: str

    temperature: str

    notes: str


class LogisticsEngine:

    def __init__(self, workbook):

        self.workbook = workbook

        self.orders = []

        self.zones = {}
        
        self.warehouse = Warehouse(workbook)

    #########################################################

    def load(self):
        
        self.orders.clear()
        self.zones.clear()
        
        
        try:

            wb = load_workbook(self.workbook)

            ws = wb["Orders"]

        except Exception as e:

            print(e)

            return False

        current_zone = None

        for row in ws.iter_rows(values_only=True):

            values = list(row)

            # Skip blank rows
            if not any(values):
                continue

            first = str(values[0]).strip() if values[0] else ""

            #######################################################
            # Detect Zone
            #######################################################

            if first.startswith("Zone"):

                try:

                    current_zone = int(first.split()[1])

                    self.zones[current_zone] = []

                    print(f"\nFound Zone {current_zone}")

                except:

                    pass

                continue

            #######################################################
            # Skip Header Row
            #######################################################

            if first == "Shipping Company":

                continue

            #######################################################
            # Skip if not inside a zone
            #######################################################

            if current_zone is None:

                continue

            #######################################################
            # Skip empty rows
            #######################################################

            if values[2] is None:

                continue

            #######################################################
            # Create Order
            #######################################################

            try:

                pallets = int(values[4]) if values[4] else 0

            except:

                pallets = 0

            order = Order(

                zone=current_zone,

                shipping_company=str(values[0] or ""),

                origin=str(values[1] or ""),

                order_number=str(values[2] or ""),

                product=str(values[3] or ""),

                pallets=pallets,

                customer=str(values[5] or ""),

                destination=str(values[6] or ""),

                warehouse_location=str(values[7] or ""),

                temperature=str(values[8] or ""),

                notes=str(values[9] or "")

            )

            self.orders.append(order)

            self.zones[current_zone].append(order)
        
        warehouse_loaded = self.warehouse.load()
        
        if not warehouse_loaded:
            print("Warning: WarehouseData could not be loaded.")
            
        return True

    #########################################################

    def find_order(self, order_number):

        for order in self.orders:

            if order.order_number == str(order_number):

                return order

        return None

    #########################################################

    def get_zone(self, zone):

        return self.zones.get(zone, [])

    #########################################################

    def find_customer(self, customer):

        results = []

        for order in self.orders:

            if order.customer.lower() == customer.lower():

                results.append(order)

        return results

    #########################################################

    def find_destination(self, destination):

        results = []

        for order in self.orders:

            if destination.lower() in order.destination.lower():

                results.append(order)

        return results

    #########################################################

    def find_temperature(self, temp):

        results = []

        for order in self.orders:

            if temp.lower() in order.temperature.lower():

                results.append(order)

        return results

    #########################################################

    def find_available_orders(self, orders):

        return self.orders
        
    #########################################################

    def score_order(self, current_order, candidate, route):

        score = 0

        ####################################################
        # Same zone group
        ####################################################

        zone_groups = {

            1: [1, 2],
            2: [1, 2],
            3: [3, 4],
            4: [3, 4]

        }

        if candidate.zone in zone_groups[current_order.zone]:
            score += 50

        ####################################################
        # Same temperature
        ####################################################

        if candidate.temperature == current_order.temperature:
            score += 40

        ####################################################
        # Same customer
        ####################################################

        if candidate.customer == current_order.customer:
            score += 20

        ####################################################
        # Same destination state
        ####################################################

        current_state = current_order.destination.split(",")[-1].strip()
        candidate_state = candidate.destination.split(",")[-1].strip()

        if current_state == candidate_state:
            score += 30

        ####################################################
        # Same product type
        ####################################################

        if current_order.product == candidate.product:
            score += 10

        ####################################################
        # Trailer capacity
        ####################################################

        if route.total_pallets + candidate.pallets > 53:

            score -= 1000

        ####################################################
        # Already on route
        ####################################################

        for stop in route.orders:

            if stop.order_number == candidate.order_number:

                score -= 1000

        return score    
    
    
    def route_planner(
            self,
            start_order,
            max_stops=5,
            max_pallets=53
        ):

        route = Route()

        # Starting order
        route.zone = start_order.zone
        route.orders.append(start_order)
        route.total_pallets = start_order.pallets

        # Zones that can work together
        zone_groups = {

            1: [1, 2],
            2: [1, 2],
            3: [3, 4],
            4: [3, 4]

        }

        allowed_zones = zone_groups.get(start_order.zone, [start_order.zone])

        # Search every order
        for order in self.orders:
            current_order = start_order

            while len(route.orders) < max_stops:

                best_order = None

                best_score = -9999

                for candidate in self.orders:

                    score = self.score_order(

                        current_order,

                        candidate,

                        route

                    )

                    if score > best_score:

                        best_score = score

                        best_order = candidate

                if best_order is None:

                    break

                route.orders.append(best_order)

                route.total_pallets += best_order.pallets

                current_order = best_order

        return route
    
    def create_route(self, order_number):

        start_order = self.find_order(order_number)

        if start_order is None:
            return None

        return self.route_planner(start_order)
        
        